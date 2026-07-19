# -*- coding: utf-8 -*-
"""Xuất tiến độ tầng C ra .xlsx — định dạng bám file PDF gốc của trung tâm.
Mở tốt trên Excel & upload thẳng lên Google Sheets (giữ màu, freeze, filter).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import _data as D

C_HEAD, C_HEAD2 = "FFFF00", "FDEADA"
C_HE, C_SOFT = "00FF00", "FFA500"
C_TEST, C_TEST_L = "29B5C4", "DCF0F3"
C_TET, C_END, C_LD = "F8CBCB", "FF0000", "FFD966"
C_CHUONG = "F2F2F2"
C_A, C_B, C_C = "C6EFCE", "FFF2CC", "E7E6E6"   # ưu tiên A / B / C

THIN = Side(style="thin", color="A6A6A6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WIDTHS = [6, 56, 24, 9, 14, 13, 9, 10, 12, 13, 17, 40, 44, 14, 12, 12, 12, 12, 14, 22]
CENTER_COLS = (1, 4, 5, 6, 7, 8, 9, 10)
DATE_FMT = "DD/MM/YYYY"

def muc_day(r):
    """Mức dạy — chỉ có nghĩa với dòng có tiết SGK. Mặc định 'Đầy đủ'."""
    if not isinstance(r[2], int) or not r[2]:
        return ""
    return r[10] if len(r) > 10 else "Đầy đủ"

def fill(c): return PatternFill("solid", fgColor=c)

def style_row(ws, r, ncol, bg=None, bold=False, color="000000", size=9):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        cell.font = Font(name="Calibri", size=size, bold=bold, color=color)
        cell.alignment = Alignment(vertical="center", wrap_text=True,
                                   horizontal="center" if c in CENTER_COLS else "left")
        if bg:
            cell.fill = fill(bg)

def banner(ws, r, text, bg, ncol, color="000000"):
    ws.cell(row=r, column=1, value=text)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    style_row(ws, r, ncol, bg=bg, bold=True, color=color, size=10)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 20

def row_bg(hm, lkt, nd):
    if nd == D.HE:                       return C_HE
    if hm in ("NT-PP", "NT-TD", "ĐL-CH", "DL"): return C_SOFT
    if lkt in ("Giữa kì", "Cuối kì", "Thi thử", "Khảo sát"): return C_TEST
    if lkt in ("15 phút", "1 tiết"):     return C_TEST_L
    return None

def build(weeks, path, title, buoi_label):
    wb = Workbook(); ws = wb.active; ws.title = title
    ncol = len(D.COLS)
    for i, h in enumerate(D.COLS, 1):
        ws.cell(row=1, column=i, value=h)
    style_row(ws, 1, ncol, bold=True)
    for i in range(1, ncol + 1):
        ws.cell(row=1, column=i).fill = fill(C_HEAD if i <= D.CORE_N else C_HEAD2)
        ws.cell(row=1, column=i).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    dates = D.sundays(max(weeks))
    r, tet_done, end_done = 2, False, False
    for w in sorted(weeks):
        d = dates[w - 1]
        # băng HẾT CHƯƠNG TRÌNH đặt ngay sau tuần cuối có nội dung SGK
        if not end_done and any("HẾT CHƯƠNG TRÌNH" in x[9] for x in weeks.get(w - 1, [])):
            banner(ws, r, f"★ HẾT CHƯƠNG TRÌNH SGK — tuần {w-1}, {dates[w-2].strftime('%d/%m/%Y')}", C_END, ncol, "FFFFFF"); r += 1
            banner(ws, r, "LUYỆN ĐỀ – ÔN THI VÀO 10  ·  18 buổi  ·  bám ma trận đề Sở GD&ĐT Hà Nội", C_LD, ncol); r += 1
            end_done = True
        if not tet_done and d > max(D.TET_OFF):
            banner(ws, r, "NGHỈ TẾT NGUYÊN ĐÁN (Đinh Mùi — mùng 1: 06/02/2027) · nghỉ 2 buổi 07/02 và 14/02", C_TET, ncol); r += 1
            tet_done = True
        r0 = r
        for i, row in enumerate(weeks[w]):
            td, bai, tiet, phut, hm, lkt, ut, cau, nd, gc = row[:10]
            vals = [w if i == 0 else None, td, bai, tiet, muc_day(row),
                    d if i == 0 else None,          # DATE THẬT, không phải chuỗi
                    phut, hm, lkt, ut, cau, nd, gc]
            for ci, v in enumerate(vals, 1):
                ws.cell(row=r, column=ci, value=v if v != "" else None)
            ws.cell(row=r, column=6).number_format = DATE_FMT
            bg = row_bg(hm, lkt, nd)
            style_row(ws, r, ncol, bg=bg)
            if not bg:
                if nd.startswith("CHƯƠNG"):
                    ws.cell(row=r, column=12).fill = fill(C_CHUONG)
                if ut:
                    ws.cell(row=r, column=10).fill = fill({D.A: C_A, D.B: C_B, D.C: C_C}[ut])
                if cau and cau != D.Q_NO:
                    ws.cell(row=r, column=11).fill = fill(C_A)
                if muc_day(row) in ("Tối thiểu", "Chỉ ý vào đề"):
                    ws.cell(row=r, column=5).fill = fill(C_C)
            ws.cell(row=r, column=2).font = Font(name="Calibri", size=9,
                                                 bold=lkt in ("Giữa kì", "Cuối kì", "Thi thử", "1 tiết"))
            r += 1
        if r - r0 > 1:
            for col in (1, 6):
                ws.merge_cells(start_row=r0, start_column=col, end_row=r - 1, end_column=col)
                ws.cell(row=r0, column=col).alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{r-1}"
    ws.sheet_view.zoomScale = 90

    lg = wb.create_sheet("Chú giải")
    lg.column_dimensions["A"].width = 24
    lg.column_dimensions["B"].width = 112
    rows = [
        ("TIẾN ĐỘ KIẾN THỨC — LỚP C", f"{title} · tầng C · năm học 2026–2027"),
        ("Nhịp buổi", f"Mỗi tuần 1 buổi vào Chủ nhật, {buoi_label}"),
        ("", ""),
        ("MỐC LỚN", ""),
        ("Tuần 1–3", "Ôn hè – nền Lớp 8 (21/06 → 05/07/2026)"),
        ("Tuần 4–30", "Học chương trình SGK — HẾT CHƯƠNG TRÌNH tuần 30 = 10/01/2027"),
        ("Tuần 31–48", "LUYỆN ĐỀ – ÔN THI VÀO 10: 18 buổi (17/01 → 30/05/2027)"),
        ("Nghỉ Tết", "07/02 và 14/02/2027 (Tết Đinh Mùi, mùng 1 = 06/02/2027)"),
        ("", ""),
        ("CỘT ƯU TIÊN", "Xếp theo MA TRẬN ĐỀ VÀO 10 HÀ NỘI — xem inputs/seeds/lop-9/de-thi-vao-10/MA-TRAN-VAO-10-HA-NOI.md"),
        ("A – Trọng tâm vào 10", "Có mặt ở 3/3 đề gốc (minh họa + 2025 + 2026). Dạy kỹ, luyện sâu."),
        ("B – Nền, cần cho thi trường", "Vào 10 không hỏi trực tiếp nhưng là nền, hoặc đề GK1/CK1 của trường hỏi nhiều."),
        ("C – Tối thiểu", "Không có trong 3/3 đề vào 10. Chỉ dạy đủ qua kiểm tra ở trường."),
        ("", ""),
        ("CỘT CÂU VÀO 10", "Buổi này phục vụ câu nào, đáng bao nhiêu điểm trong đề 10 điểm."),
        ("", ""),
        ("CỘT TIẾT SGK (PPCT)", "Số tiết mà PPCT của SGV phân cho bài đó — là SỐ THAM CHIẾU, KHÔNG phải cam kết dạy đủ."),
        ("CỘT MỨC DẠY", "Đầy đủ = dạy trọn bài. Chỉ ý vào đề = chỉ luyện ý có trong đề, bỏ phần còn lại. Tối thiểu = chỉ đủ qua kiểm tra ở trường."),
        ("", ""),
        ("MA TRẬN CHÍNH THỨC CỦA SỞ", "Số & Đại số 4,5đ · Hình học và Đo lường 4,0đ · Thống kê và Xác suất 1,5đ"),
        ("theo cấp độ tư duy", "Nhận biết 2,75đ · Thông hiểu 4,25đ · Vận dụng 3,0đ  →  NB+TH = 7,0đ = trần điểm thực tế của tầng C"),
        ("Thống kê–Xác suất", "1,0đ NB + 0,5đ TH, KHÔNG có câu vận dụng → 1,5đ dễ ăn nhất đề, ưu tiên số 1 cho lớp C"),
        ("", ""),
        ("HẠNG MỤC", ""),
        ("KT", "Kiến thức — dạy bài mới / luyện tập / chữa đề"),
        ("Test", "Kiểm tra (xem cột Loại kiểm tra)"),
        ("NT-PP / NT-TD", "Nhận thức – Phương pháp học / Tư duy"),
        ("ĐL-CH / DL", "Động lực – Câu hỏi / Động lực"),
        ("", ""),
        ("LOẠI KIỂM TRA", ""),
        ("15 phút", "Đầu buổi — soi bài + BTVN buổi trước (bỏ ở buổi mở chương)"),
        ("1 tiết", "45′ cuối chương trọng tâm (III, V, VI, IX), kèm 45′ chữa đề ngay trong buổi"),
        ("Giữa kì / Cuối kì", "90′ — lần 2–4 giao về nhà, luôn có buổi chữa đề"),
        ("Thi thử", "Đề thi thử vào 10, thi chung toàn hệ thống"),
        ("", ""),
        ("MÀU NỀN", ""),
        ("Xanh lá / Cam", "Khối ôn hè / Dòng phương pháp học – động lực"),
        ("Xanh ngọc đậm / nhạt", "Kiểm tra lớn (GK, CK, thi thử) / Kiểm tra 15 phút, 1 tiết"),
        ("Đỏ / Hồng / Vàng", "Băng mốc: hết chương trình · nghỉ Tết · vào pha luyện đề"),
        ("", ""),
        ("NGUỒN", "Tên bài + số tiết: PPCT trong SGV Toán 9 KNTT (tr.17–19) + mục lục SGK tập 1 & 2. Ma trận: 3 đề gốc của Sở GD&ĐT Hà Nội (đã lưu ở inputs/seeds/lop-9/de-thi-vao-10/)."),
        ("ĐÃ BỎ", "10 tiết Hoạt động thực hành trải nghiệm của SGK (GeoGebra, pha chế dung dịch, gene trội) — không nằm trong nội dung thi. Thầy chốt 16/07/2026."),
        ("ĐÃ GIẢM MẠNH", "Tỉ số lượng giác Ch IV (11 tiết → 2 buổi, chỉ đủ GK1 trường) · BĐT–BPT Bài 5–6 · đồ thị y=ax² Bài 18 · đa giác đều Bài 30 · vị trí 2 đường tròn Bài 17 · độ dài cung Bài 15. Tất cả đều KHÔNG có ở 3/3 đề vào 10."),
        ("ĐÃ TRẢ GIỜ VỀ", "Thống kê–Xác suất (Câu I, 1,5đ) và Hình khối (Câu IV.1, ~1,0đ) — mỗi phần 3 buổi theo Thầy chốt, chỉ luyện đúng ý vào đề."),
        ("CỘT SẢN XUẤT", "Trạng thái · Deadline L0–L4 · Người sản xuất · ND chiếu màn hình — để trống, điền dần khi giao việc."),
    ]
    for i, (a, b) in enumerate(rows, 1):
        lg.cell(row=i, column=1, value=a).font = Font(bold=True, size=10)
        lg.cell(row=i, column=2, value=b).font = Font(size=10)
        lg.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        if a and not b:
            lg.cell(row=i, column=1).fill = fill(C_HEAD)
    lg.cell(row=1, column=1).font = Font(bold=True, size=12)
    for k, c in [("A – Trọng tâm vào 10", C_A), ("B – Nền, cần cho thi trường", C_B), ("C – Tối thiểu", C_C)]:
        for i, (a, _) in enumerate(rows, 1):
            if a == k:
                lg.cell(row=i, column=1).fill = fill(c)

    wb.save(path)
    print(f"✓ {path}  ({r-2} dòng)")


BASE = "/Users/admin/Documents/thaitd/Code/giaoanMathtech/inputs/seeds/lop-9/chuong-trinh-hoc/"
build(D.DAI_SO, BASE + "[C]dai-so-tien-do-kien-thuc.xlsx", "Lớp 9 · Đại số",
      "180′ (165′ học + 15′ giải lao) — mỗi buổi = 1 phiếu")
build(D.HINH_HOC, BASE + "[C]hinh-hoc-tien-do-kien-thuc.xlsx", "Lớp 9 · Hình học",
      "90′ — mỗi buổi = 1 phiếu")
